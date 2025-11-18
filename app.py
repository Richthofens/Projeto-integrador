from flask import Flask, request, render_template, url_for, redirect, make_response
from config import get_db_connection
import csv
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import date


app = Flask(__name__)


def get_dashboard_data():
    db = get_db_connection()
    cursor = db.cursor()

    # Total de EPIs
    cursor.execute("SELECT COUNT(*) FROM epi")
    total_epi = cursor.fetchone()[0]

    # Total de funcionários
    cursor.execute("SELECT COUNT(*) FROM users")
    total_func = cursor.fetchone()[0]

    # Funcionários treinados no último ano
    cursor.execute(
        "SELECT COUNT(*) FROM users WHERE data_treinamento IS NOT NULL AND data_treinamento >= DATE_SUB(CURDATE(), INTERVAL 365 DAY)"
    )
    total_trained = cursor.fetchone()[0]

    # EPIs vencendo nos próximos 30 dias
    cursor.execute(
        "SELECT COUNT(*) FROM epi WHERE dataValidade BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY)"
    )
    epivencendo_count = cursor.fetchone()[0]

    # Lista detalhada de EPIs vencendo (nome, dataValidade)
    cursor.execute(
        "SELECT nome, dataValidade FROM epi WHERE dataValidade BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 30 DAY) ORDER BY dataValidade ASC"
    )
    soon_expiring = cursor.fetchall()

    # Percentual de conformidade de treinamento
    training_percentage = None
    if total_func > 0:
        training_percentage = round((total_trained / total_func) * 100, 1)

    # Todos os EPIs (para tabela)
    cursor.execute("SELECT id, nome, codigo, lote, quantidadeTotal, dataValidade FROM epi ORDER BY nome")
    all_epis = cursor.fetchall()

    # Todos os funcionários (para tabela)
    cursor.execute("SELECT id, nome, cargo, setor, epi_atribuido, data_treinamento FROM users ORDER BY nome")
    all_funcionarios = cursor.fetchall()

    db.close()

    return {
        'total_epi': total_epi,
        'total_func': total_func,
        'total_trained': total_trained,
        'epivencendo_count': epivencendo_count,
        'soon_expiring': soon_expiring,
        'training_percentage': training_percentage,
        'reports_generated': None,
        'epis': all_epis,
        'funcionarios': all_funcionarios,
        'now': date.today(),
    }


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        form = request.form

        # Detecta formulário de Funcionário por um campo único (ex.: 'cargo')
        # Caso contrário assume que é o formulário de EPI
        if form.get('cargo') is not None:
            try:
                db = get_db_connection()
                cursor = db.cursor()

                cursor.execute(
                    "INSERT INTO users (nome, cargo, setor, epi_atribuido, data_treinamento) VALUES (%s, %s, %s, %s, %s)",
                    (form.get('nome'), form.get('cargo'), form.get('setor'), form.get('epis'), form.get('treinamento'))
                )
                db.commit()
                db.close()

            except Exception as e:
                print('Erro ao inserir Funcionário:', e)

        else:
            try:
                db = get_db_connection()
                cursor = db.cursor()
                validade = form.get('validade') or '2025-12-31'

                cursor.execute(
                    "INSERT INTO epi (nome, codigo, lote, dataValidade, quantidadeTotal, fornecedor) VALUES (%s, %s, %s, %s, %s, %s)",
                    (form.get('nome'), form.get('codigo'), form.get('lote'), validade, int(form.get('quantidade', 0)), form.get('fornecedor'))
                )

                db.commit()
                db.close()

            except Exception as e:
                print('Erro ao inserir EPI:', e)

        return redirect(url_for('index'))

    data = get_dashboard_data()
    return render_template('epi360.html', **data)


@app.route('/deletar_epi/<int:epi_id>', methods=['POST'])
def deletar_epi(epi_id):
    # Verifica senha de admin
    senha = request.form.get('senha', '')
    senha_admin = 'admin123'  # Mude isso para uma variável de ambiente em produção
    
    if senha != senha_admin:
        return {'status': 'erro', 'mensagem': 'Senha incorreta'}, 401
    
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("DELETE FROM epi WHERE id = %s", (epi_id,))
        db.commit()
        db.close()
        return {'status': 'sucesso', 'mensagem': 'EPI deletado com sucesso'}, 200
    except Exception as e:
        return {'status': 'erro', 'mensagem': str(e)}, 500


@app.route('/deletar_funcionario/<int:user_id>', methods=['POST'])
def deletar_funcionario(user_id):
    # Verifica senha de admin
    senha = request.form.get('senha', '')
    senha_admin = 'admin123'  # Mude isso para uma variável de ambiente em produção
    
    if senha != senha_admin:
        return {'status': 'erro', 'mensagem': 'Senha incorreta'}, 401
    
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        db.commit()
        db.close()
        return {'status': 'sucesso', 'mensagem': 'Funcionário deletado com sucesso'}, 200
    except Exception as e:
        return {'status': 'erro', 'mensagem': str(e)}, 500


@app.route('/gerar_relatorio', methods=['GET'])
def gerar_relatorio():
    from datetime import datetime, date

    # Função auxiliar para converter datas antes de exportar
    def excel_safe_date(value):
        if isinstance(value, (datetime, date)):
            return value
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except:
            return None  # evita números como 46073 virarem datas erradas

    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("SELECT nome, codigo, lote, dataValidade, quantidadeTotal, fornecedor FROM epi ORDER BY nome")
    epis = cursor.fetchall()

    cursor.execute("SELECT nome, cargo, setor, epi_atribuido, data_treinamento FROM users ORDER BY nome")
    users = cursor.fetchall()

    db.close()

    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Títulos mesclados
    ws.merge_cells('A1:F1')
    ws['A1'] = 'EPIs'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True)
    for col in ['A','B','C','D','E','F']:
        ws[f"{col}1"].border = border

    ws.merge_cells('H1:L1')
    ws['H1'] = 'FUNCIONARIOS'
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].font = Font(bold=True)
    for col in ['H','I','J','K','L']:
        ws[f"{col}1"].border = border

    # Headers
    epi_headers = ['nome', 'codigo', 'lote', 'dataValidade', 'quantidadeTotal', 'fornecedor']
    user_headers = ['nome', 'cargo', 'setor', 'epi_atribuido', 'data_treinamento']

    for idx, h in enumerate(epi_headers, start=1):
        cell = ws.cell(row=2, column=idx, value=h)
        cell.font = Font(bold=True)
        cell.border = border

    for jdx, h in enumerate(user_headers, start=8):
        cell = ws.cell(row=2, column=jdx, value=h)
        cell.font = Font(bold=True)
        cell.border = border

    # Popula dados
    max_linhas = max(len(epis), len(users))

    for i in range(max_linhas):
        row_num = 3 + i

        # EPIs
        if i < len(epis):
            for c_idx, value in enumerate(epis[i], start=1):
                cell = ws.cell(row=row_num, column=c_idx)

                # Coluna 4 = dataValidade
                if c_idx == 4:
                    value = excel_safe_date(value)
                    cell.number_format = "dd/mm/yyyy"

                cell.value = value
        else:
            for c_idx in range(1, 7):
                ws.cell(row=row_num, column=c_idx, value='')

        # FUNCIONÁRIOS
        if i < len(users):
            for u_idx, value in enumerate(users[i], start=8):
                cell = ws.cell(row=row_num, column=u_idx)

                # Coluna 12 = data_treinamento
                if u_idx == 12:
                    value = excel_safe_date(value)
                    cell.number_format = "dd/mm/yyyy"

                cell.value = value
        else:
            for u_idx in range(8, 13):
                ws.cell(row=row_num, column=u_idx, value='')

    # Ajusta largura
    for col in ['A','B','C','D','E','F','H','I','J','K','L']:
        ws.column_dimensions[col].width = 18

    # Salva em memória
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = make_response(bio.read())
    response.headers['Content-Disposition'] = 'attachment; filename=relatorio_epi360.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


if __name__ == '__main__':
    app.run(debug=True)
